import streamlit as st
import pandas as pd
import numpy as np
import datetime
import io
from supabase import create_client, Client
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

# --- PRODUCTION AUTHENTICATION ---
SUPABASE_URL = "https://wljftpkvsozgpxivbwiu.supabase.co"
SUPABASE_KEY = "sb_secret_60ve-Yh8xAvI6MZkhQQR3Q_Fk2mP9If"
supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

st.set_page_config(page_title="GLOBO Master Analytics Engine", layout="wide")

st.title("📊 GLOBO Master Report & Advanced Export Ledger")
st.subheader("⚡ High Performance Server-Cached Database Edition")

def fetch_prebuilt_ledger(start_iso, end_iso):
    """Fetches transactional data rows comprehensively from master_reporting_ledger."""
    all_rows = []
    chunk_size = 1000
    start_idx = 0
    
    while True:
        res = supabase.table("master_reporting_ledger") \
                      .select("*") \
                      .gte("Created at", start_iso) \
                      .lte("Created at", end_iso) \
                      .range(start_idx, start_idx + chunk_size - 1) \
                      .execute()
        
        data_chunk = res.data
        if not data_chunk:
            break
        all_rows.extend(data_chunk)
        if len(data_chunk) < chunk_size:
            break
        start_idx += chunk_size
        
    df = pd.DataFrame(all_rows)
    if not df.empty:
        df["created_at_dt"] = pd.to_datetime(df["Created at"], errors='coerce').dt.tz_localize(None)
    return df

# Columns mapped out strictly based on your specified sample layout template
EXPORT_COLUMNS = [
    "Serial No", "Name", "SR Order ID", "Created at", "Financial Status", "Fulfillment Status",
    "Currency", "Subtotal", "Shipping", "Taxes", "Total", "Shipping Method",
    "Outstanding Balance", "Tax 1 Name", "Tax 1 Value", "Billing Province Name",
    "Shipping Province Name", "Payment Mode", "Lineitem name", "Lineitem quantity",
    "Lineitem price", "HSN CODE", "SHOPIFY DELIVERY STATUS", "awb number", "SHIPROCKET DELIVERY STATUS"
]

MASTER_COLS = [
    "Serial No", "Name", "Created at", "Financial Status", "Fulfillment Status", 
    "Currency", "Subtotal", "Shipping", "Taxes", "Total", "Shipping Method", 
    "Outstanding Balance", "Billing Province Name", "Shipping Province Name", "Payment Mode"
]

def generate_excel_with_merged_cells(raw_df):
    """
    Generates a professionally styled binary .xlsx file where all shared master order level 
    columns are physically merged into tall visual cells using openpyxl.
    """
    # Sort data layout sequentially to bundle families
    df = raw_df.sort_values(by=["Created at", "shopify_lineitem_id", "shiprocket_shipment_id"], ascending=[False, True, True]).copy()
    
    unique_orders = df["Name"].unique()
    order_to_serial = {name: idx + 1 for idx, name in enumerate(unique_orders)}
    df["Serial No"] = df["Name"].map(order_to_serial)
    
    for col in EXPORT_COLUMNS:
        if col not in df.columns:
            df[col] = np.nan
            
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Order Report"
    
    # 1. Styling Definitions
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=11)
    
    thin_side = Side(border_style="thin", color="D9D9D9")
    cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # Write Headers
    ws.append(EXPORT_COLUMNS)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
    
    # Track grouping clusters to trigger physical mergers
    order_groups = []
    current_order = None
    start_row = 2
    
    for idx, row in df.iterrows():
        row_data = [None if pd.isna(row.get(c)) else row.get(c) for c in EXPORT_COLUMNS]
        ws.append(row_data)
        current_row_idx = ws.max_row
        
        # Apply style to added row cells
        for col_idx in range(1, len(EXPORT_COLUMNS) + 1):
            cell = ws.cell(row=current_row_idx, column=col_idx)
            cell.font = data_font
            cell.border = cell_border
            # Choose clean left alignment for labels, center for tracking metrics/codes
            if EXPORT_COLUMNS[col_idx-1] in ["Name", "Lineitem name"]:
                cell.alignment = left_align
            else:
                cell.alignment = center_align
                
        # Calculate row boundaries for order clusters
        order_name = row.get("Name")
        if current_order is None:
            current_order = order_name
            start_row = current_row_idx
        elif order_name != current_order:
            order_groups.append((start_row, current_row_idx - 1))
            current_order = order_name
            start_row = current_row_idx
            
    if current_order is not None:
        order_groups.append((start_row, ws.max_row))
        
    # Execute physical merges cell-by-cell over calculated index matrices
    for s_row, e_row in order_groups:
        if s_row == e_row:
            continue
        for col_name in MASTER_COLS:
            col_idx = EXPORT_COLUMNS.index(col_name) + 1
            ws.merge_cells(start_row=s_row, start_column=col_idx, end_row=e_row, end_column=col_idx)
            
            # Re-ensure top-level vertical centering is preserved over newly merged configurations
            master_cell = ws.cell(row=s_row, column=col_idx)
            master_cell.alignment = center_align

    # Auto-fit structural column sizing widths neatly
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

def render_html_merged_dashboard(raw_df):
    """
    Compiles standard data matrices directly into an HTML data table, using 
    rowspan attributes to beautifully render merged blocks natively on the screen.
    """
    df = raw_df.sort_values(by=["Created at", "shopify_lineitem_id", "shiprocket_shipment_id"], ascending=[False, True, True]).copy()
    unique_orders = df["Name"].unique()
    order_to_serial = {name: idx + 1 for idx, name in enumerate(unique_orders)}
    df["Serial No"] = df["Name"].map(order_to_serial)
    
    # Calculate exact lifespan counts per order to set structural rowspans
    order_counts = df["Name"].value_counts().to_dict()
    seen_orders = set()
    
    html = """
    <style>
        .merged-table { width: 100%; border-collapse: collapse; font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; font-size: 13px; }
        .merged-table th { background-color: #1F4E78; color: white; padding: 10px; border: 1px solid #D9D9D9; position: sticky; top: 0; }
        .merged-table td { padding: 8px; border: 1px solid #E2E8F0; vertical-align: middle; text-align: center; background-color: #FFFFFF; }
        .merged-table .left-align { text-align: left; }
        .merged-table tr:hover td { background-color: #F8FAFC !important; }
    </style>
    <div style="overflow-x: auto; max-height: 600px; border: 1px solid #CBD5E1; border-radius: 6px;">
    <table class="merged-table">
        <thead>
            <tr>
    """
    for col in EXPORT_COLUMNS:
        html += f"<th>{col}</th>"
    html += "</tr></thead><tbody>"
    
    for _, row in df.iterrows():
        html += "tr"
        order_name = row.get("Name")
        row_span = order_counts.get(order_name, 1)
        
        is_first_row_of_cluster = order_name not in seen_orders
        
        for col in EXPORT_COLUMNS:
            val = row.get(col)
            val_str = "" if pd.isna(val) else str(val)
            
            # Format high-precision financial metrics cleanly
            if col in ["Subtotal", "Shipping", "Taxes", "Total", "Outstanding Balance", "Tax 1 Value", "Lineitem price"] and val_str:
                try: val_str = f"₹{float(val_str):,.2f}"
                except: pass
                
            if col in MASTER_COLS:
                if is_first_row_of_cluster:
                    # Inject rowspan attribute to merge cells vertically on your dashboard screen
                    align_class = " class='left-align'" if col == "Name" else ""
                    html += f"<td rowspan='{row_span}'{align_class}>{val_str}</td>"
            else:
                align_class = " class='left-align'" if col == "Lineitem name" else ""
                html += f"<td{align_class}>{val_str}</td>"
                
        html += "</tr>"
        seen_orders.add(order_name)
        
    html += "</tbody></table></div>"
    return html

# --- SIDEBAR COMPACT TIMELINE REGULATORS ---
st.sidebar.header("📅 Query Range Controller")
filter_type = st.sidebar.radio("Mode:", ["Exact Calendar Dates", "Whole Month / Year"])

if filter_type == "Exact Calendar Dates":
    today = datetime.date.today()
    start_date = st.sidebar.date_input("Start Date", today - datetime.timedelta(days=30))
    end_date = st.sidebar.date_input("End Date", today)
    start_iso = f"{start_date}T00:00:00Z"
    end_iso = f"{end_date}T23:59:59Z"
else:
    current_year = datetime.date.today().year
    year = st.sidebar.selectbox("Year", list(range(current_year, current_year - 4, -1)))
    months_map = {
        "January": 1, "February": 2, "March": 3, "April": 4, "May": 5, "June": 6,
        "July": 7, "August": 8, "September": 9, "October": 10, "November": 11, "December": 12
    }
    month_name = st.sidebar.selectbox("Month", list(months_map.keys()))
    month_num = months_map[month_name]
    
    start_iso = f"{year}-{month_num:02d}-01T00:00:00Z"
    if month_num == 12:
        end_iso = f"{year}-12-31T23:59:59Z"
    else:
        end_iso = f"{(datetime.date(year, month_num + 1, 1) - datetime.timedelta(days=1)).isoformat()}T23:59:59Z"

# --- RUN EXECUTION PIPELINE ---
with st.spinner("⚡ Querying database targets..."):
    raw_df = fetch_prebuilt_ledger(start_iso, end_iso)

if raw_df.empty:
    st.warning("ℹ️ No records found matching this date slice.")
else:
    # Compile key totals safely
    unique_orders_count = raw_df["Name"].nunique()
    total_rev = pd.to_numeric(raw_df.drop_duplicates(subset=["Name"])["Total"], errors='coerce').sum()
    unique_awbs_count = raw_df["awb number"].dropna().nunique()

    # --- SEARCH BAR ---
    search_query = st.text_input("🔍 Search within filtered results (Order Name, AWB, Province)", "")
    if search_query:
        raw_df = raw_df[
            raw_df["Name"].astype(str).str.contains(search_query, case=False, na=False) |
            raw_df["awb number"].astype(str).str.contains(search_query, case=False, na=False) |
            raw_df["Shipping Province Name"].astype(str).str.contains(search_query, case=False, na=False)
        ]

    # --- METRICS DISPLAYS ---
    st.markdown("---")
    c1, c2, c3 = st.columns(3)
    c1.metric("Unique Orders Found", unique_orders_count)
    c2.metric("Total Period Revenue", f"₹{total_rev:,.2f}")
    c3.metric("Packages Tracked", unique_awbs_count)

    # --- EXPORT INTERACTIVE INTERFACE LINK ---
    st.sidebar.markdown("---")
    with st.spinner("📦 Pre-rendering Excel row-merges..."):
        excel_binary_data = generate_excel_with_merged_cells(raw_df)
        
    st.sidebar.download_button(
        label="📥 Download Merged Excel (.xlsx)",
        data=excel_binary_data,
        file_name=f"GLOBO_Merged_Report_{start_iso[:10]}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    # --- RENDER WEB COMPILING INTERFACE MATRIX ---
    st.markdown("### 📋 Structured Data Matrix (Merged Row View)")
    dashboard_html_table = render_html_merged_dashboard(raw_df)
    st.markdown(dashboard_html_table, unsafe_allow_html=True)
